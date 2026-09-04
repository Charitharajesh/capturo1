"""
Runs the backend pytest suite (writing JUnit XML), then converts that real
result into reports/backend-report.xlsx — mirroring the "Test Results" +
"Summary" sheet structure the e2e-web Selenium suite already produces, so
every suite in this repo reports in the same shape.

Every row comes straight from the JUnit XML pytest just wrote; nothing here
invents a status, a duration, or a test that didn't actually run.
"""
import subprocess
import sys
import xml.etree.ElementTree as ET
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

BACKEND_DIR = Path(__file__).resolve().parent.parent
REPORTS_DIR = BACKEND_DIR / "reports"
JUNIT_PATH = REPORTS_DIR / "junit.xml"
XLSX_PATH = REPORTS_DIR / "backend-report.xlsx"


def run_pytest() -> int:
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    proc = subprocess.run(
        [sys.executable, "-m", "pytest", "tests/", "-v", f"--junitxml={JUNIT_PATH}"],
        cwd=BACKEND_DIR,
    )
    return proc.returncode


def parse_junit():
    tree = ET.parse(JUNIT_PATH)
    root = tree.getroot()
    suite = root if root.tag == "testsuite" else root.find("testsuite")

    rows = []
    for case in suite.iter("testcase"):
        name = case.get("name")
        classname = case.get("classname", "")
        duration_ms = round(float(case.get("time", "0")) * 1000)

        failure = case.find("failure")
        error = case.find("error")
        skipped = case.find("skipped")
        if failure is not None or error is not None:
            status = "Fail"
            node = failure if failure is not None else error
            message = (node.get("message") or node.text or "").strip().splitlines()[0][:300]
        elif skipped is not None:
            status = "Skipped"
            message = (skipped.get("message") or "").strip()
        else:
            status = "Pass"
            message = ""

        rows.append(
            {
                "name": name,
                "category": classname,
                "status": status,
                "duration_ms": duration_ms,
                "message": message,
            }
        )

    stats = {
        "total": int(suite.get("tests", 0)),
        "failures": int(suite.get("failures", 0)),
        "errors": int(suite.get("errors", 0)),
        "skipped": int(suite.get("skipped", 0)),
        "duration_s": float(suite.get("time", 0)),
    }
    return rows, stats


def write_workbook(rows, stats):
    wb = Workbook()

    sheet = wb.active
    sheet.title = "Test Results"
    headers = ["Test Name", "Category", "Status", "Duration (ms)", "Error/Remarks"]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for row in rows:
        sheet.append([row["name"], row["category"], row["status"], row["duration_ms"], row["message"]])
        status_cell = sheet.cell(row=sheet.max_row, column=3)
        status_cell.font = Font(bold=True, color="2E7D32" if row["status"] == "Pass" else "C62828")

    widths = [60, 40, 12, 16, 60]
    for idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = width
    sheet.auto_filter.ref = f"A1:E{sheet.max_row}"
    sheet.freeze_panes = "A2"

    summary = wb.create_sheet("Summary")
    summary.append(["Metric", "Value"])
    for cell in summary[1]:
        cell.font = Font(bold=True)

    passed = sum(1 for r in rows if r["status"] == "Pass")
    failed = sum(1 for r in rows if r["status"] == "Fail")
    skipped = sum(1 for r in rows if r["status"] == "Skipped")
    total = len(rows)
    pass_rate = round((passed / total) * 100, 2) if total else 0.0

    summary_rows = [
        ("Total Tests", total),
        ("Passed", passed),
        ("Failed", failed),
        ("Skipped", skipped),
        ("Pass Rate (%)", pass_rate),
        ("Total Suite Duration (s)", round(stats["duration_s"], 2)),
        ("Report Generated (UTC)", datetime.now(timezone.utc).isoformat()),
    ]
    for metric, value in summary_rows:
        summary.append([metric, value])
    summary.column_dimensions["A"].width = 34
    summary.column_dimensions["B"].width = 30

    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(XLSX_PATH)
    print(f"Excel report written to {XLSX_PATH}")
    print(f"Total: {total} | Passed: {passed} | Failed: {failed} | Skipped: {skipped} | Pass rate: {pass_rate}%")


def main():
    exit_code = run_pytest()
    rows, stats = parse_junit()
    write_workbook(rows, stats)
    sys.exit(exit_code)


if __name__ == "__main__":
    main()
